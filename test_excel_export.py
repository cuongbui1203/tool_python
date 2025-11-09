#!/usr/bin/env python3
"""
Test script để kiểm tra export Excel
"""

from csv_processor_v2 import CSVProcessorV2, Config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def test_export():
    # Tạo config
    config = Config(
        parametric_name_column="parametric",
        get_columns=["min", "max"],
        begin_from_parametric=False,
        null_values=["N/A", "NULL", "-"],
        key_column="key"
    )
    
    # So sánh files
    result = CSVProcessorV2.process_files("test_dummy.csv", "dummy2.csv", config)
    
    print(f"New params: {len(result.new_params)}")
    print(f"Removed params: {len(result.removed_params)}")
    print(f"Changed params: {len(result.changed_params)}")
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()

    ws.title = "Comparison Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # LEFT SIDE: Summary Info
    ws['A1'] = 'SW Version'
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws['A1'].fill = header_fill
    ws['A1'].border = thin_border
    
    ws['B1'] = 'Total keys'
    ws['B1'].font = header_font
    ws['B1'].alignment = center_alignment
    ws['B1'].fill = header_fill
    ws['B1'].border = thin_border
    
    ws['A2'] = result.old_version
    ws['A2'].alignment = center_alignment
    ws['A2'].border = thin_border
    
    ws['B2'] = result.total_old_version
    ws['B2'].alignment = center_alignment
    ws['B2'].border = thin_border
    
    ws['A3'] = result.new_version
    ws['A3'].alignment = center_alignment
    ws['A3'].border = thin_border
    
    ws['B3'] = result.total_new_version
    ws['B3'].alignment = center_alignment
    ws['B3'].border = thin_border
    
    # Type of change table
    row = 5
    ws[f'A{row}'] = 'Type of change'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center_alignment
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = 'Quantity'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].alignment = center_alignment
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].border = thin_border
    
    row += 1
    ws[f'A{row}'] = 'Added Keys'
    ws[f'A{row}'].alignment = left_alignment
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'] = len(result.new_params)
    ws[f'B{row}'].alignment = center_alignment
    ws[f'B{row}'].border = thin_border
    
    row += 1
    ws[f'A{row}'] = 'Removed Keys'
    ws[f'A{row}'].alignment = left_alignment
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'] = len(result.removed_params)
    ws[f'B{row}'].alignment = center_alignment
    ws[f'B{row}'].border = thin_border
    
    # RIGHT SIDE
    ws['D1'] = f'Bundle {result.old_version.replace(".csv", "")} VS bundle {result.new_version.replace(".csv", "")}'
    ws.merge_cells('D1:I1')
    ws['D1'].font = Font(bold=True, size=12)
    ws['D1'].alignment = center_alignment
    ws['D1'].border = thin_border
    
    # Headers
    ws['E7'] = 'Key Name'
    ws['E7'].font = header_font
    ws['E7'].alignment = center_alignment
    ws['E7'].fill = header_fill
    ws['E7'].border = thin_border
    
    ws['F7'] = 'UL'
    ws['F7'].font = header_font
    ws['F7'].alignment = center_alignment
    ws['F7'].fill = header_fill
    ws['F7'].border = thin_border
    
    ws['G7'] = 'LL'
    ws['G7'].font = header_font
    ws['G7'].alignment = center_alignment
    ws['G7'].fill = header_fill
    ws['G7'].border = thin_border
    
    # Data
    current_row = 8
    
    # Removed Keys
    for param in result.removed_params:
        ws[f'D{current_row}'] = 'Removed Keys'
        ws[f'D{current_row}'].fill = red_fill
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws[f'E{current_row}'] = param.name
        ws[f'E{current_row}'].border = thin_border
        
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['E'].width = 30
    
    # Save
    wb.save('test_export.xlsx')
    print("✅ Exported to test_export.xlsx")

if __name__ == "__main__":
    test_export()
