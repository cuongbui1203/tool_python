import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import tkinter.font as tkfont
from typing import Optional
import threading
import os
import pandas as pd
from datetime import datetime
from csv_processor_v2 import CSVProcessorV2, ComparisonResult, Config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
            

class MaterialColors:
    """Material Design color palette"""

    PRIMARY = "#1976D2"
    PRIMARY_DARK = "#1565C0"
    PRIMARY_LIGHT = "#42A5F5"
    SECONDARY = "#FF5722"
    BACKGROUND = "#FAFAFA"
    SURFACE = "#FFFFFF"
    ERROR = "#F44336"
    SUCCESS = "#4CAF50"
    WARNING = "#FF9800"
    TEXT_PRIMARY = "#212121"
    TEXT_SECONDARY = "#757575"
    DIVIDER = "#E0E0E0"


class MaterialButton(tk.Button):
    """Material Design styled button"""

    def __init__(self, parent, text="", command=None, style="primary", **kwargs):
        # Material button colors
        colors = {
            "primary": {
                "bg": MaterialColors.PRIMARY,
                "fg": "#212121",
                "active_bg": MaterialColors.PRIMARY_DARK,
            },
            "secondary": {
                "bg": MaterialColors.SECONDARY,
                "fg": "#212121",
                "active_bg": "#E64A19",
            },
            "success": {
                "bg": MaterialColors.SUCCESS,
                "fg": "#212121",
                "active_bg": "#388E3C",
            },
            "outline": {
                "bg": MaterialColors.SURFACE,
                "fg": MaterialColors.PRIMARY,
                "active_bg": MaterialColors.BACKGROUND,
            },
        }

        color_scheme = colors.get(style, colors["primary"])

        super().__init__(
            parent,
            text=text,
            command=command,  # type: ignore
            bg=color_scheme["bg"],
            fg=color_scheme["fg"],
            activebackground=color_scheme["active_bg"],
            activeforeground="white",
            relief="flat",
            borderwidth=0,
            padx=20,
            pady=10,
            font=("Segoe UI", 10, "bold"),
            cursor="hand2",
            **kwargs,
        )

        # Hover effects
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.default_bg = color_scheme["bg"]
        self.hover_bg = color_scheme["active_bg"]

    def _on_enter(self, event):
        self.config(bg=self.hover_bg)

    def _on_leave(self, event):
        self.config(bg=self.default_bg)


class MaterialCard(tk.Frame):
    """Material Design card component"""

    def __init__(self, parent, title="", **kwargs):
        super().__init__(
            parent,
            bg=MaterialColors.SURFACE,
            relief="flat",
            borderwidth=1,
            highlightbackground=MaterialColors.DIVIDER,
            highlightthickness=1,
            **kwargs,
        )

        if title:
            title_label = tk.Label(
                self,
                text=title,
                bg=MaterialColors.SURFACE,
                fg=MaterialColors.TEXT_PRIMARY,
                font=("Segoe UI", 12, "bold"),
                anchor="w",
            )
            title_label.pack(fill="x", padx=16, pady=(16, 8))


class CSVComparatorGUI:
    """GUI chính cho CSV Comparator với Material Design"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("D03 Burn-in Parametric Data Comparison Tool")
        self.root.geometry("1200x800")
        self.root.configure(bg=MaterialColors.BACKGROUND)

        # Force light mode for all OS
        self.force_light_mode()

        # Variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.comparison_result: Optional[ComparisonResult] = None
        
        # Config variables
        self.parametric_column = tk.StringVar(value="parametric")
        self.get_columns = tk.StringVar(value="min,max")
        self.null_values = tk.StringVar(value="N/A,NULL,-")
        self.key_column = tk.StringVar(value="key")
        self.begin_from_parametric = tk.BooleanVar(value=False)

        self.setup_styles()
        self.create_widgets()

    def force_light_mode(self):
        """Force light mode theme for all operating systems"""
        try:
            # For Windows - disable dark mode
            if hasattr(self.root, 'tk_setPalette'):
                self.root.tk_setPalette(
                    background=MaterialColors.BACKGROUND,
                    foreground=MaterialColors.TEXT_PRIMARY,
                    activeBackground=MaterialColors.PRIMARY_LIGHT,
                    activeForeground="white"
                )
            
            # Set window attributes for consistent appearance
            self.root.option_add('*TkFDialog*foreground', MaterialColors.TEXT_PRIMARY)
            self.root.option_add('*TkFDialog*background', MaterialColors.SURFACE)
            
            # Force light theme for ttk widgets
            style = ttk.Style()
            
            # Available themes - prefer light themes
            available_themes = style.theme_names()
            light_themes = ['clam', 'alt', 'default', 'classic']
            
            selected_theme = 'clam'  # Default fallback
            for theme in light_themes:
                if theme in available_themes:
                    selected_theme = theme
                    break
            
            style.theme_use(selected_theme)
            
            # Override system dark mode settings
            self.root.tk.call('tk_setPalette', 
                             MaterialColors.BACKGROUND,
                             foreground=MaterialColors.TEXT_PRIMARY, # type: ignore
                             activeBackground=MaterialColors.PRIMARY_LIGHT, # type: ignore
                             activeForeground="white", # type: ignore
                             selectBackground=MaterialColors.PRIMARY_LIGHT, # type: ignore
                             selectForeground="white") # type: ignore
                             
        except Exception as e:
            print(f"Warning: Could not force light mode: {e}")

    def setup_styles(self):
        """Thiết lập styles cho ttk widgets"""
        style = ttk.Style()

        # Force light theme colors for all widgets
        style.configure('.',
                       background=MaterialColors.SURFACE,
                       foreground=MaterialColors.TEXT_PRIMARY,
                       fieldbackground=MaterialColors.SURFACE,
                       selectbackground=MaterialColors.PRIMARY_LIGHT,
                       selectforeground="white")

        # Configure Notebook style
        style.configure(
            "Material.TNotebook", 
            background=MaterialColors.BACKGROUND, 
            borderwidth=0,
            tabmargins=[0, 5, 0, 0]
        )
        style.configure(
            "Material.TNotebook.Tab",
            background=MaterialColors.SURFACE,
            foreground=MaterialColors.TEXT_PRIMARY,
            padding=[20, 10],
            font=("Liberation Sans", 10),
            borderwidth=1,
            focuscolor='none'
        )
        style.map(
            "Material.TNotebook.Tab",
            background=[("selected", MaterialColors.PRIMARY), ("active", MaterialColors.PRIMARY_LIGHT)],
            foreground=[("selected", "white"), ("active", "white")],
            bordercolor=[("selected", MaterialColors.PRIMARY), ("active", MaterialColors.PRIMARY_LIGHT)]
        )

                # Configure Treeview style for tables
        style.configure(
            "Material.Treeview",
            background=MaterialColors.SURFACE,
            foreground=MaterialColors.TEXT_PRIMARY,
            fieldbackground=MaterialColors.SURFACE,
            borderwidth=0,
            font=("Liberation Sans", 10),
            rowheight=25
        )
        style.configure(
            "Material.Treeview.Heading",
            background=MaterialColors.PRIMARY,
            foreground="white",
            font=("Liberation Sans", 10, "bold"),
            borderwidth=1,
            relief="flat",
        )
        style.map(
            "Material.Treeview",
            background=[("selected", MaterialColors.PRIMARY_LIGHT)],
            foreground=[("selected", "white")],
        )
        style.map(
            "Material.Treeview.Heading",
            background=[("active", MaterialColors.PRIMARY_DARK)],
            foreground=[("active", "white")]
        )

        # Configure Scrollbar style
        style.configure("Material.Vertical.TScrollbar",
                       background=MaterialColors.SURFACE,
                       troughcolor=MaterialColors.BACKGROUND,
                       bordercolor=MaterialColors.DIVIDER,
                       arrowcolor=MaterialColors.TEXT_SECONDARY,
                       darkcolor=MaterialColors.SURFACE,
                       lightcolor=MaterialColors.SURFACE)
        
        style.configure("Material.Horizontal.TScrollbar",
                       background=MaterialColors.SURFACE,
                       troughcolor=MaterialColors.BACKGROUND,
                       bordercolor=MaterialColors.DIVIDER,
                       arrowcolor=MaterialColors.TEXT_SECONDARY,
                       darkcolor=MaterialColors.SURFACE,
                       lightcolor=MaterialColors.SURFACE)

    def create_widgets(self):
        """Tạo giao diện chính"""
        # Header
        self.create_header()

        # Main content
        main_frame = tk.Frame(self.root, bg=MaterialColors.BACKGROUND)
        main_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # File selection section
        self.create_file_selection(main_frame)

        # Results section
        self.create_results_section(main_frame)

    def create_header(self):
        """Tạo header của ứng dụng"""
        header_frame = tk.Frame(self.root, bg=MaterialColors.PRIMARY, height=80)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        title_label = tk.Label(
            header_frame,
            text="D03 Burn-in Parametric Data Comparison",
            bg=MaterialColors.PRIMARY,
            fg="white",
            font=("Arial", 20, "bold"),
        )
        title_label.pack(expand=True)

        subtitle_label = tk.Label(
            header_frame,
            text="Compare parametric data between two CSV files",
            bg=MaterialColors.PRIMARY,
            fg="white",
            font=("Segoe UI", 10),
        )
        subtitle_label.pack()

    def create_file_selection(self, parent):
        """Tạo phần chọn file"""
        file_card = MaterialCard(parent, title="Select Files to Compare")
        file_card.pack(fill="x", pady=(0, 20))

        # File 1
        file1_frame = tk.Frame(file_card, bg=MaterialColors.SURFACE)
        file1_frame.pack(fill="x", padx=16, pady=8)

        tk.Label(
            file1_frame,
            text="File 1 (Old Version):",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        file1_input_frame = tk.Frame(file1_frame, bg=MaterialColors.SURFACE)
        file1_input_frame.pack(fill="x", pady=(5, 0))

        self.file1_entry = tk.Entry(
            file1_input_frame,
            textvariable=self.file1_path,
            font=("Liberation Sans", 10),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            insertbackground=MaterialColors.TEXT_PRIMARY,
            selectbackground=MaterialColors.PRIMARY_LIGHT,
            selectforeground="white"
        )
        self.file1_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        MaterialButton(
            file1_input_frame,
            text="Browse",
            command=lambda: self.browse_file(self.file1_path),
            style="secondary",
        ).pack(side="right")

        # File 2
        file2_frame = tk.Frame(file_card, bg=MaterialColors.SURFACE)
        file2_frame.pack(fill="x", padx=16, pady=8)

        tk.Label(
            file2_frame,
            text="File 2 (New Version):",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        file2_input_frame = tk.Frame(file2_frame, bg=MaterialColors.SURFACE)
        file2_input_frame.pack(fill="x", pady=(5, 0))

        self.file2_entry = tk.Entry(
            file2_input_frame,
            textvariable=self.file2_path,
            font=("Liberation Sans", 10),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            insertbackground=MaterialColors.TEXT_PRIMARY,
            selectbackground=MaterialColors.PRIMARY_LIGHT,
            selectforeground="white"
        )
        self.file2_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        MaterialButton(
            file2_input_frame,
            text="Browse",
            command=lambda: self.browse_file(self.file2_path),
            style="secondary",
        ).pack(side="right")

        # Advanced Configuration Section
        self.create_config_section(file_card)

        # Compare button
        button_frame = tk.Frame(file_card, bg=MaterialColors.SURFACE)
        button_frame.pack(fill="x", padx=16, pady=(8, 16))

        self.compare_button = MaterialButton(
            button_frame,
            text="Compare Files",
            command=self.compare_files,
            style="primary",
        )
        self.compare_button.pack(pady=10)

        # Export Excel button
        self.export_button = MaterialButton(
            button_frame,
            text="📊 Export to Excel",
            command=self.export_to_excel,
            style="success",
        )
        self.export_button.pack(pady=(5, 10))
        self.export_button.config(state="disabled")  # Disabled until comparison is done

        # Status label
        self.status_label = tk.Label(
            button_frame,
            text="Select two CSV files to start comparison",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_SECONDARY,
            font=("Liberation Sans", 9),
        )
        self.status_label.pack()
    
    def create_config_section(self, parent):
        """Tạo phần cấu hình nâng cao"""
        # Collapsible config section
        config_frame = tk.Frame(parent, bg=MaterialColors.SURFACE)
        config_frame.pack(fill="x", padx=16, pady=(8, 0))
        
        # Toggle button for config
        self.config_visible = tk.BooleanVar(value=False)
        toggle_btn = tk.Button(
            config_frame,
            text="⚙️ Advanced Configuration (Click to expand)",
            command=self.toggle_config,
            bg=MaterialColors.BACKGROUND,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            cursor="hand2",
            anchor="w",
            padx=10,
            pady=5
        )
        toggle_btn.pack(fill="x")
        
        # Config content frame (initially hidden)
        self.config_content = tk.Frame(config_frame, bg=MaterialColors.SURFACE)
        
        # Parametric Column Name
        param_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        param_frame.pack(fill="x", pady=5)
        
        tk.Label(
            param_frame,
            text="Parametric Column Name:",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Liberation Sans", 9),
            width=25,
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            param_frame,
            textvariable=self.parametric_column,
            font=("Liberation Sans", 9),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            width=30
        ).pack(side="left", padx=5)
        
        # Get Columns
        cols_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        cols_frame.pack(fill="x", pady=5)
        
        tk.Label(
            cols_frame,
            text="Columns to Extract (comma-separated):",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Liberation Sans", 9),
            width=25,
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            cols_frame,
            textvariable=self.get_columns,
            font=("Liberation Sans", 9),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            width=30
        ).pack(side="left", padx=5)
        
        # Null Values
        null_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        null_frame.pack(fill="x", pady=5)
        
        tk.Label(
            null_frame,
            text="Null Values (comma-separated):",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Liberation Sans", 9),
            width=25,
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            null_frame,
            textvariable=self.null_values,
            font=("Liberation Sans", 9),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            width=30
        ).pack(side="left", padx=5)
        
        # Key Column
        key_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        key_frame.pack(fill="x", pady=5)
        
        tk.Label(
            key_frame,
            text="Key Column Name:",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Liberation Sans", 9),
            width=25,
            anchor="w"
        ).pack(side="left")
        
        tk.Entry(
            key_frame,
            textvariable=self.key_column,
            font=("Liberation Sans", 9),
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            relief="solid",
            borderwidth=1,
            width=30
        ).pack(side="left", padx=5)
        
        # Begin from Parametric checkbox
        begin_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        begin_frame.pack(fill="x", pady=5)
        
        tk.Checkbutton(
            begin_frame,
            text="Begin from Parametric Column",
            variable=self.begin_from_parametric,
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Liberation Sans", 9),
            selectcolor=MaterialColors.SURFACE,
            activebackground=MaterialColors.SURFACE,
            activeforeground=MaterialColors.PRIMARY
        ).pack(side="left")
        
        # Reset button
        reset_btn_frame = tk.Frame(self.config_content, bg=MaterialColors.SURFACE)
        reset_btn_frame.pack(fill="x", pady=5)
        
        MaterialButton(
            reset_btn_frame,
            text="Reset to Default",
            command=self.reset_config,
            style="outline"
        ).pack(side="left")
    
    def toggle_config(self):
        """Toggle hiển thị config section"""
        if self.config_visible.get():
            self.config_content.pack_forget()
            self.config_visible.set(False)
        else:
            self.config_content.pack(fill="x", pady=5)
            self.config_visible.set(True)
    
    def reset_config(self):
        """Reset config về giá trị mặc định"""
        self.parametric_column.set("parametric")
        self.get_columns.set("min,max")
        self.null_values.set("N/A,NULL,-")
        self.key_column.set("key")
        self.begin_from_parametric.set(False)
        messagebox.showinfo("Reset", "Configuration reset to default values")

    def create_results_section(self, parent):
        """Tạo phần hiển thị kết quả"""
        results_card = MaterialCard(parent, title="📊 Comparison Results")
        results_card.pack(fill="both", expand=True)

        # Notebook for tabs
        self.notebook = ttk.Notebook(results_card, style="Material.TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=16, pady=(8, 16))

        # Create tabs
        self.create_summary_tab()
        self.create_new_params_tab()
        self.create_removed_params_tab()
        self.create_changed_params_tab()

    def create_summary_tab(self):
        """Tạo tab tổng quan với thống kê dạng bảng"""
        frame = tk.Frame(self.notebook, bg=MaterialColors.SURFACE)
        self.notebook.add(frame, text="📈 Overview")

        # Stats summary frame
        stats_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        stats_frame.pack(fill="x", padx=10, pady=(10, 5))

        # Files info
        files_label = tk.Label(
            stats_frame,
            text="📁 Compared Files:",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        files_label.pack(fill="x")

        self.files_info_label = tk.Label(
            stats_frame,
            text="",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_SECONDARY,
            font=("Segoe UI", 9),
            anchor="w",
            wraplength=1000,
        )
        self.files_info_label.pack(fill="x", padx=20)

        # Stats table
        stats_table_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        stats_table_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(
            stats_table_frame,
            text="Overview Statistics:",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        ).pack(fill="x")

        # Create stats table
        self.stats_table = ttk.Treeview(
            stats_table_frame,
            columns=("category", "count"),
            show="headings",
            height=4,
            style="Material.Treeview",
        )

        # Configure columns
        self.stats_table.heading("category", text="Change Type")
        self.stats_table.heading("count", text="Quantity")
        # self.stats_table.heading("percentage", text="Tỷ lệ (%)")

        self.stats_table.column("category", width=200, anchor="center")
        self.stats_table.column("count", width=100, anchor="center")
        # self.stats_table.column("percentage", width=100, anchor="center")

        self.stats_table.pack(fill="x", pady=5)

        # Summary text for additional details
        detail_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        detail_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # tk.Label(
        #     detail_frame,
        #     text="🔍 Chi tiết nhanh:",
        #     bg=MaterialColors.SURFACE,
        #     fg=MaterialColors.TEXT_PRIMARY,
        #     font=("Segoe UI", 12, "bold"),
        #     anchor="w"
        # ).pack(fill="x")

        # self.summary_text = scrolledtext.ScrolledText(
        #     detail_frame,
        #     wrap=tk.WORD,
        #     bg="white",
        #     fg=MaterialColors.TEXT_PRIMARY,
        #     font=("Segoe UI", 9),
        #     relief="flat",
        #     borderwidth=1,
        #     height=8
        # )
        # self.summary_text.pack(fill="both", expand=True, pady=5)

    def create_new_params_tab(self):
        """Tạo tab parameters mới với bảng"""
        frame = tk.Frame(self.notebook, bg=MaterialColors.SURFACE)
        self.notebook.add(frame, text="🆕 New Parametric Keys")

        # Header
        header_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        header_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.new_params_count_label = tk.Label(
            header_frame,
            text="🆕 New Parametric Keys",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        self.new_params_count_label.pack(fill="x")

        # Create table
        table_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Treeview with scrollbar
        tree_frame = tk.Frame(table_frame, bg=MaterialColors.SURFACE)
        tree_frame.pack(fill="both", expand=True)

        self.new_params_table = ttk.Treeview(
            tree_frame,
            columns=("name", "upper_limit", "lower_limit", "status"),
            show="headings",
            style="Material.Treeview",
        )

        # Configure columns
        self.new_params_table.heading("name", text="Parameter Name")
        self.new_params_table.heading("upper_limit", text="Upper Limit")
        self.new_params_table.heading("lower_limit", text="Lower Limit")
        self.new_params_table.heading("status", text="Status")

        self.new_params_table.column("name", width=250, anchor="w")
        self.new_params_table.column("upper_limit", width=120, anchor="center")
        self.new_params_table.column("lower_limit", width=120, anchor="center")
        self.new_params_table.column("status", width=100, anchor="center")

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.new_params_table.yview,
            style="Material.Vertical.TScrollbar"
        )
        h_scrollbar = ttk.Scrollbar(
            tree_frame, orient="horizontal", command=self.new_params_table.xview,
            style="Material.Horizontal.TScrollbar"
        )

        self.new_params_table.configure(
            yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set
        )

        # Pack table and scrollbars
        self.new_params_table.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

    def create_removed_params_tab(self):
        """Tạo tab parameters bị xóa với bảng"""
        frame = tk.Frame(self.notebook, bg=MaterialColors.SURFACE)
        self.notebook.add(frame, text="❌ Removed Parametric Keys")

        # Header
        header_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        header_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.removed_params_count_label = tk.Label(
            header_frame,
            text="❌ Removed Parametric Keys",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        self.removed_params_count_label.pack(fill="x")

        # Create table
        table_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree_frame = tk.Frame(table_frame, bg=MaterialColors.SURFACE)
        tree_frame.pack(fill="both", expand=True)

        self.removed_params_table = ttk.Treeview(
            tree_frame,
            columns=("name", "upper_limit", "lower_limit", "status"),
            show="headings",
            style="Material.Treeview",
        )

        # Configure columns
        self.removed_params_table.heading("name", text="Parameter Name")
        self.removed_params_table.heading("upper_limit", text="Upper Limit")
        self.removed_params_table.heading("lower_limit", text="Lower Limit")
        self.removed_params_table.heading("status", text="Status")

        self.removed_params_table.column("name", width=250, anchor="w")
        self.removed_params_table.column("upper_limit", width=120, anchor="center")
        self.removed_params_table.column("lower_limit", width=120, anchor="center")
        self.removed_params_table.column("status", width=100, anchor="center")

        # Scrollbars
        v_scrollbar2 = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.removed_params_table.yview,
            style="Material.Vertical.TScrollbar"
        )
        self.removed_params_table.configure(yscrollcommand=v_scrollbar2.set)

        self.removed_params_table.pack(side="left", fill="both", expand=True)
        v_scrollbar2.pack(side="right", fill="y")

    def create_changed_params_tab(self):
        """Tạo tab parameters thay đổi với bảng so sánh"""
        frame = tk.Frame(self.notebook, bg=MaterialColors.SURFACE)
        self.notebook.add(frame, text="🔄 Changed Parametric Keys")

        # Header
        header_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        header_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.changed_params_count_label = tk.Label(
            header_frame,
            text="🔄 Changed Parametric Keys",
            bg=MaterialColors.SURFACE,
            fg=MaterialColors.TEXT_PRIMARY,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        self.changed_params_count_label.pack(fill="x")

        # Create table
        table_frame = tk.Frame(frame, bg=MaterialColors.SURFACE)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree_frame = tk.Frame(table_frame, bg=MaterialColors.SURFACE)
        tree_frame.pack(fill="both", expand=True)

        self.changed_params_table = ttk.Treeview(
            tree_frame,
            columns=(
                "name",
                "old_upper",
                "old_lower",
                "new_upper",
                "new_lower",
                "change_type",
            ),
            show="headings",
            style="Material.Treeview",
        )

        # Configure columns
        self.changed_params_table.heading("name", text="Parameter Name")
        self.changed_params_table.heading("old_upper", text="Old Upper")
        self.changed_params_table.heading("old_lower", text="Old Lower")
        self.changed_params_table.heading("new_upper", text="New Upper")
        self.changed_params_table.heading("new_lower", text="New Lower")
        self.changed_params_table.heading("change_type", text="Change Type")

        self.changed_params_table.column("name", width=200, anchor="w")
        self.changed_params_table.column("old_upper", width=100, anchor="center")
        self.changed_params_table.column("old_lower", width=100, anchor="center")
        self.changed_params_table.column("new_upper", width=100, anchor="center")
        self.changed_params_table.column("new_lower", width=100, anchor="center")
        self.changed_params_table.column("change_type", width=120, anchor="center")

        # Scrollbars
        v_scrollbar3 = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.changed_params_table.yview,
            style="Material.Vertical.TScrollbar"
        )
        h_scrollbar3 = ttk.Scrollbar(
            tree_frame, orient="horizontal", command=self.changed_params_table.xview,
            style="Material.Horizontal.TScrollbar"
        )

        self.changed_params_table.configure(
            yscrollcommand=v_scrollbar3.set, xscrollcommand=h_scrollbar3.set
        )

        self.changed_params_table.pack(side="left", fill="both", expand=True)
        v_scrollbar3.pack(side="right", fill="y")

    def browse_file(self, var: tk.StringVar):
        """Mở dialog chọn file"""
        filename = filedialog.askopenfilename(
            title="Chọn CSV file",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if filename:
            var.set(filename)

    def compare_files(self):
        """So sánh 2 files"""
        file1 = self.file1_path.get().strip()
        file2 = self.file2_path.get().strip()

        if not file1 or not file2:
            messagebox.showerror("Error", "Please select two CSV files!")
            return

        if not os.path.exists(file1):
            messagebox.showerror("Error", f"File not exist: {file1}")
            return

        if not os.path.exists(file2):
            messagebox.showerror("Error", f"File not exist: {file2}")
            return

        # Disable button và hiển thị loading
        self.compare_button.config(state="disabled", text="⏳ Comparing...")
        self.status_label.config(
            text="Processing data...", fg=MaterialColors.WARNING
        )

        # Chạy so sánh trong thread riêng
        thread = threading.Thread(target=self.perform_comparison, args=(file1, file2))
        thread.daemon = True
        thread.start()

    def perform_comparison(self, file1: str, file2: str):
        """Thực hiện so sánh trong background thread"""
        try:
            # Tạo config từ UI
            config = Config(
                parametric_name_column=self.parametric_column.get(),
                get_columns=self.get_columns.get().split(','),
                begin_from_parametric=self.begin_from_parametric.get(),
                null_values=self.null_values.get().split(','),
                key_column=self.key_column.get()
            )
            
            # So sánh files với config
            result = CSVProcessorV2.process_files(file1, file2, config)

            # Cập nhật UI trong main thread
            self.root.after(0, self.update_results, result, None)

        except Exception as e:
            # Cập nhật UI với lỗi
            self.root.after(0, self.update_results, None, str(e))

    def update_results(self, result: Optional[ComparisonResult], error: Optional[str]):
        """Cập nhật kết quả lên UI"""
        # Enable button
        self.compare_button.config(state="normal", text="🔍 Compare Files")

        if error:
            self.status_label.config(text=f"Error: {error}", fg=MaterialColors.ERROR)
            self.export_button.config(state="disabled")
            messagebox.showerror("Error", error)
            return

        if not result:
            return

        self.comparison_result = result
        self.status_label.config(text="Comparison completed!", fg=MaterialColors.SUCCESS)
        
        # Enable export button
        self.export_button.config(state="normal")

        # Clear previous results
        self.clear_results()

        # Update summary
        self.update_summary_tab(result)

        # Update other tabs
        self.update_new_params_tab(result.new_params)
        self.update_removed_params_tab(result.removed_params)
        self.update_changed_params_tab(result.changed_params)

        # Switch to summary tab
        self.notebook.select(0)

    def clear_results(self):
        """Xóa kết quả cũ"""
        # self.summary_text.delete(1.0, tk.END)

        # Clear tables
        for item in self.new_params_table.get_children():
            self.new_params_table.delete(item)
        for item in self.removed_params_table.get_children():
            self.removed_params_table.delete(item)
        for item in self.changed_params_table.get_children():
            self.changed_params_table.delete(item)
        for item in self.stats_table.get_children():
            self.stats_table.delete(item)

    def update_summary_tab(self, result: ComparisonResult):
        """Cập nhật tab tổng quan với bảng thống kê"""
        # Update files info
        files_info = f"• File 1 (Old Version): {self.file1_path.get()}\n• File 2 (New Version): {self.file2_path.get()}"
        self.files_info_label.config(text=files_info)

        # Calculate totals
        total_changes = (
            len(result.new_params)
            + len(result.removed_params)
            + len(result.changed_params)
        )

        # Update stats table
        stats_data = [
            (
                "🆕 New Parametric Keys",
                len(result.new_params),
            ),
            (
                "❌ Removed Parametric Keys",
                len(result.removed_params),
            ),
            (
                "🔄 Changed Parametric Keys",
                len(result.changed_params),
            ),
            ("🔄 Total Change", total_changes),
            (f"Qty Parametric Keys of Bundle: {result.old_version}", result.total_old_version),
            (f"Qty Parametric Keys of Bundle: {result.new_version}", result.total_new_version),
        ]

        for category, count in stats_data:
            self.stats_table.insert("", "end", values=(category, count))

        # Update summary text with quick details
        summary = "🔍 QUICK DETAILS:\n\n"

        if result.new_params:
            summary += f"🆕 TOP NEW PARAMETRIC KEYS:\n"
            for i, param in enumerate(result.new_params[:5], 1):  # Show first 5
                summary += f"   {i}. {param.name}\n"
            if len(result.new_params) > 5:
                summary += f"   ... và {len(result.new_params) - 5} parameters khác\n"

        if result.removed_params:
            summary += f"\n❌ TOP REMOVED PARAMETRIC KEYS:\n"
            for i, param in enumerate(result.removed_params[:5], 1):  # Show first 5
                summary += f"   {i}. {param.name}\n"
            if len(result.removed_params) > 5:
                summary += (
                    f"   ... và {len(result.removed_params) - 5} parameters khác\n"
                )

        if result.changed_params:
            summary += f"\n🔄 TOP CHANGED PARAMETRIC KEYS:\n"
            for i, change in enumerate(result.changed_params[:5], 1):  # Show first 5
                summary += f"   {i}. {change.old.name}\n"
            if len(result.changed_params) > 5:
                summary += (
                    f"   ... và {len(result.changed_params) - 5} parameters khác\n"
                )

        if total_changes == 0:
            summary += "✅ No differences found between the two files!"

        # self.summary_text.insert(1.0, summary)

    def update_new_params_tab(self, new_params):
        """Cập nhật tab parameters mới với bảng"""
        # Update header
        count_text = f"🆕 New Parametric Keys ({len(new_params)} items)"
        self.new_params_count_label.config(text=count_text)

        if not new_params:
            # Insert empty message
            self.new_params_table.insert(
                "",
                "end",
                values=("No New Parametric Keys", "-", "-", "No changes"),
            )
            return

        # Populate table
        for i, param in enumerate(new_params, 1):
            upper_val = (
                param.limit.upper_limit
                if param.limit.upper_limit is not None
                else "N/A"
            )
            lower_val = (
                param.limit.lower_limit
                if param.limit.lower_limit is not None
                else "N/A"
            )

            self.new_params_table.insert(
                "",
                "end",
                values=(param.name, upper_val, lower_val, "NEW"),
                tags=("new",),
            )

        # Configure tag colors
        self.new_params_table.tag_configure(
            "new", background="#E8F5E8", foreground="#2E7D32"
        )

    def update_removed_params_tab(self, removed_params):
        """Cập nhật tab parameters bị xóa với bảng"""
        # Update header
        count_text = f"❌ Removed Parametric Keys ({len(removed_params)} items)"
        self.removed_params_count_label.config(text=count_text)

        if not removed_params:
            self.removed_params_table.insert(
                "",
                "end",
                values=("No Parametric Keys were deleted", "-", "-", "No changes"),
            )
            return

        # Populate table
        for i, param in enumerate(removed_params, 1):
            upper_val = (
                param.limit.upper_limit
                if param.limit.upper_limit is not None
                else "N/A"
            )
            lower_val = (
                param.limit.lower_limit
                if param.limit.lower_limit is not None
                else "N/A"
            )

            self.removed_params_table.insert(
                "",
                "end",
                values=(param.name, upper_val, lower_val, "REMOVED"),
                tags=("removed",),
            )

        # Configure tag colors
        self.removed_params_table.tag_configure(
            "removed", background="#FFEBEE", foreground="#C62828"
        )

    def update_changed_params_tab(self, changed_params):
        """Cập nhật tab parameters thay đổi với bảng so sánh"""
        # Update header
        count_text = f"🔄 CHANGED PARAMETRIC KEYS ({len(changed_params)} items)"
        self.changed_params_count_label.config(text=count_text)

        if not changed_params:
            self.changed_params_table.insert(
                "",
                "end",
                values=(
                    "Không có parameters nào thay đổi",
                    "-",
                    "-",
                    "-",
                    "-",
                    "No changes",
                ),
            )
            return

        # Populate table
        for i, change in enumerate(changed_params, 1):
            old_upper = (
                change.old.limit.upper_limit
                if change.old.limit.upper_limit is not None
                else "N/A"
            )
            old_lower = (
                change.old.limit.lower_limit
                if change.old.limit.lower_limit is not None
                else "N/A"
            )
            new_upper = (
                change.new.limit.upper_limit
                if change.new.limit.upper_limit is not None
                else "N/A"
            )
            new_lower = (
                change.new.limit.lower_limit
                if change.new.limit.lower_limit is not None
                else "N/A"
            )

            # Determine change type
            change_type = []
            if old_upper != new_upper:
                change_type.append("Upper")
            if old_lower != new_lower:
                change_type.append("Lower")
            change_type_str = " + ".join(change_type) if change_type else "Other"

            self.changed_params_table.insert(
                "",
                "end",
                values=(
                    change.old.name,
                    old_upper,
                    old_lower,
                    new_upper,
                    new_lower,
                    change_type_str,
                ),
                tags=("changed",),
            )

        # Configure tag colors
        self.changed_params_table.tag_configure(
            "changed", background="#FFF3E0", foreground="#E65100"
        )

    def export_to_excel(self):
        """Export comparison results to Excel file"""
        if not self.comparison_result:
            messagebox.showwarning("Warning", "No data to export. Please compare the files first.")
            return

        try:
            # Ask user where to save
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"CSV_Comparison_{timestamp}.xlsx"
            
            filepath = filedialog.asksaveasfilename(
                title="Save comparison results",
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filepath:
                return
            
            # Show progress
            self.export_button.config(state="disabled", text="⏳ Exporting...")
            self.status_label.config(text="Creating Excel file...", fg=MaterialColors.WARNING)
            print(f"Exporting to {filepath}")
            # Run export in thread
            thread = threading.Thread(target=self.perform_excel_export, args=(filepath,))
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            print(f"Error occurred: {e}")
            messagebox.showerror("Error", f"Error exporting Excel file: {str(e)}")
            self.export_button.config(state="normal", text="📊 Export to Excel")

    def perform_excel_export(self, filepath: str):
        """Perform Excel export in background thread with custom layout"""
        try:
            result = self.comparison_result
            if not result:
                raise ValueError("No comparison result to export.")
            
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet()

            ws.title = "Comparison Report"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # ===== LEFT SIDE: Summary Info =====
            row = 1
            
            # SW Version info
            ws['A1'] = 'SW Version'
            ws['A1'].font = Font(bold=True)
            ws['A1'].alignment = center_alignment
            ws['A1'].fill = header_fill
            ws['A1'].font = header_font
            ws['A1'].border = thin_border
            
            ws['B1'] = 'Total keys'
            ws['B1'].font = header_font
            ws['B1'].alignment = center_alignment
            ws['B1'].fill = header_fill
            ws['B1'].border = thin_border
            
            # Old version
            ws['A2'] = result.old_version
            ws['A2'].alignment = center_alignment
            ws['A2'].border = thin_border
            ws['B2'] = result.total_old_version
            ws['B2'].alignment = center_alignment
            ws['B2'].border = thin_border
            
            # New version
            ws['A3'] = result.new_version
            ws['A3'].alignment = center_alignment
            ws['A3'].border = thin_border
            ws['B3'] = result.total_new_version
            ws['B3'].alignment = center_alignment
            ws['B3'].border = thin_border
            
            # Type of change table
            row = 5
            ws[f'A{row}'] = 'Type of change'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = 'Quantity'
            ws[f'B{row}'].font = header_font
            ws[f'B{row}'].alignment = center_alignment
            ws[f'B{row}'].fill = header_fill
            ws[f'B{row}'].border = thin_border
            
            row += 1
            ws[f'A{row}'] = 'Added Keys'
            ws[f'A{row}'].alignment = left_alignment
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'] = len(result.new_params)
            ws[f'B{row}'].alignment = center_alignment
            ws[f'B{row}'].border = thin_border
            
            row += 1
            ws[f'A{row}'] = 'Removed Keys'
            ws[f'A{row}'].alignment = left_alignment
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'] = len(result.removed_params)
            ws[f'B{row}'].alignment = center_alignment
            ws[f'B{row}'].border = thin_border
            
            row += 1
            ws[f'A{row}'] = 'Limits Changed Keys'
            ws[f'A{row}'].alignment = left_alignment
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'] = len(result.changed_params)
            ws[f'B{row}'].alignment = center_alignment
            ws[f'B{row}'].border = thin_border
            
            row += 1
            ws[f'A{row}'] = 'Overlap Keys'
            ws[f'A{row}'].alignment = left_alignment
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'] = len(result.overlap_params)  # Calculate if needed
            ws[f'B{row}'].alignment = center_alignment
            ws[f'B{row}'].border = thin_border
            
            # ===== RIGHT SIDE: Comparison Details =====
            # Header row 1
            ws['D1'] = f'Bundle {result.old_version.replace(".csv", "")} VS bundle {result.new_version.replace(".csv", "")}'
            ws.merge_cells('D1:I1')
            ws['D1'].font = Font(bold=True, size=12)
            ws['D1'].alignment = center_alignment
            ws['D1'].border = thin_border
            
            # Product, Build, HW Info, SW Ver (rows 2-6)
            ws['D2'] = 'Product'
            ws['D2'].border = thin_border
            ws['E2'] = 'Jxx'
            ws['E2'].border = thin_border
            ws['F2'] = ''
            ws['F2'].border = thin_border
            ws['G2'] = ''
            ws['G2'].border = thin_border
            ws['H2'] = 'Jxx'
            ws['H2'].border = thin_border
            ws['I2'] = ''
            ws['I2'].border = thin_border
            
            ws['D3'] = 'Build'
            ws['D3'].border = thin_border
            ws['E3'] = ''
            ws['E3'].border = thin_border
            ws['F3'] = ''
            ws['F3'].border = thin_border
            ws['G3'] = ''
            ws['G3'].border = thin_border
            ws['H3'] = ''
            ws['H3'].border = thin_border
            ws['I3'] = ''
            ws['I3'].border = thin_border
            
            ws['D4'] = 'HW Info'
            ws['D4'].border = thin_border
            ws['E4'] = 'Config type'
            ws['E4'].border = thin_border
            ws['F4'] = ''
            ws['F4'].border = thin_border
            ws['G4'] = ''
            ws['G4'].border = thin_border
            ws['H4'] = ''
            ws['H4'].border = thin_border
            ws['I4'] = ''
            ws['I4'].border = thin_border
            
            ws['D5'] = ''
            ws['D5'].border = thin_border
            ws['E5'] = 'SN'
            ws['E5'].border = thin_border
            ws['F5'] = ''
            ws['F5'].border = thin_border
            ws['G5'] = ''
            ws['G5'].border = thin_border
            ws['H5'] = ''
            ws['H5'].border = thin_border
            ws['I5'] = ''
            ws['I5'].border = thin_border
            
            ws['D6'] = 'SW Ver'
            ws['D6'].border = thin_border
            ws['E6'] = result.old_version.replace(".csv", "")
            ws['E6'].border = thin_border
            ws['F6'] = ''
            ws['F6'].border = thin_border
            ws['G6'] = ''
            ws['G6'].border = thin_border
            ws['H6'] = result.new_version.replace(".csv", "")
            ws['H6'].border = thin_border
            ws['I6'] = ''
            ws['I6'].border = thin_border
            
            # Result header (row 7)
            ws['D7'] = 'Result'
            ws['D7'].border = thin_border
            ws['E7'] = 'Key Name'
            ws['E7'].font = Font(bold=True)
            ws['E7'].alignment = center_alignment
            ws['E7'].fill = header_fill
            ws['E7'].font = header_font
            ws['E7'].border = thin_border
            
            ws['F7'] = 'UL'
            ws['F7'].font = header_font
            ws['F7'].alignment = center_alignment
            ws['F7'].fill = header_fill
            ws['F7'].border = thin_border
            
            ws['G7'] = 'LL'
            ws['G7'].font = header_font
            ws['G7'].alignment = center_alignment
            ws['G7'].fill = header_fill
            ws['G7'].border = thin_border
            
            ws['H7'] = 'UL'
            ws['H7'].font = header_font
            ws['H7'].alignment = center_alignment
            ws['H7'].fill = header_fill
            ws['H7'].border = thin_border
            
            ws['I7'] = 'LL'
            ws['I7'].font = header_font
            ws['I7'].alignment = center_alignment
            ws['I7'].fill = header_fill
            ws['I7'].border = thin_border
            
            # Data rows
            current_row = 8
            
            # Removed Keys (Red)
            for param in result.removed_params:
                ws[f'D{current_row}'] = 'Removed Keys'
                ws[f'D{current_row}'].fill = red_fill
                ws[f'D{current_row}'].alignment = center_alignment
                ws[f'D{current_row}'].border = thin_border
                
                ws[f'E{current_row}'] = param.name
                ws[f'E{current_row}'].border = thin_border
                
                # Get all limit values from the data dict
                old_ul = param.limit.data.get('max') or param.limit.data.get('upper', 'NA')
                old_ll = param.limit.data.get('min') or param.limit.data.get('lower', 'NA')
                
                ws[f'F{current_row}'] = old_ul if old_ul != 'NA' else 'NA'
                ws[f'F{current_row}'].alignment = center_alignment
                ws[f'F{current_row}'].border = thin_border
                
                ws[f'G{current_row}'] = old_ll if old_ll != 'NA' else 'NA'
                ws[f'G{current_row}'].alignment = center_alignment
                ws[f'G{current_row}'].border = thin_border
                
                ws[f'H{current_row}'] = '/'
                ws[f'H{current_row}'].alignment = center_alignment
                ws[f'H{current_row}'].border = thin_border
                
                ws[f'I{current_row}'] = '/'
                ws[f'I{current_row}'].alignment = center_alignment
                ws[f'I{current_row}'].border = thin_border
                
                current_row += 1
            
            # Changed Keys (Yellow)
            for change in result.changed_params:
                ws[f'D{current_row}'] = 'Limits Changed Keys'
                ws[f'D{current_row}'].fill = yellow_fill
                ws[f'D{current_row}'].alignment = center_alignment
                ws[f'D{current_row}'].border = thin_border
                
                ws[f'E{current_row}'] = change.old.name
                ws[f'E{current_row}'].border = thin_border
                
                old_ul = change.old.limit.data.get('max') or change.old.limit.data.get('upper', 'NA')
                old_ll = change.old.limit.data.get('min') or change.old.limit.data.get('lower', 'NA')
                new_ul = change.new.limit.data.get('max') or change.new.limit.data.get('upper', 'NA')
                new_ll = change.new.limit.data.get('min') or change.new.limit.data.get('lower', 'NA')
                
                ws[f'F{current_row}'] = old_ul if old_ul != 'NA' else 'NA'
                ws[f'F{current_row}'].alignment = center_alignment
                ws[f'F{current_row}'].border = thin_border
                
                ws[f'G{current_row}'] = old_ll if old_ll != 'NA' else 'NA'
                ws[f'G{current_row}'].alignment = center_alignment
                ws[f'G{current_row}'].border = thin_border
                
                ws[f'H{current_row}'] = new_ul if new_ul != 'NA' else 'NA'
                ws[f'H{current_row}'].alignment = center_alignment
                ws[f'H{current_row}'].border = thin_border
                
                ws[f'I{current_row}'] = new_ll if new_ll != 'NA' else 'NA'
                ws[f'I{current_row}'].alignment = center_alignment
                ws[f'I{current_row}'].border = thin_border
                
                current_row += 1
            
            # Added Keys (Blue)
            for param in result.new_params:
                ws[f'D{current_row}'] = 'Added Keys'
                ws[f'D{current_row}'].fill = blue_fill
                ws[f'D{current_row}'].alignment = center_alignment
                ws[f'D{current_row}'].border = thin_border
                
                ws[f'E{current_row}'] = param.name
                ws[f'E{current_row}'].border = thin_border
                
                ws[f'F{current_row}'] = '/'
                ws[f'F{current_row}'].alignment = center_alignment
                ws[f'F{current_row}'].border = thin_border
                
                ws[f'G{current_row}'] = '/'
                ws[f'G{current_row}'].alignment = center_alignment
                ws[f'G{current_row}'].border = thin_border
                
                new_ul = param.limit.data.get('max') or param.limit.data.get('upper', 'NA')
                new_ll = param.limit.data.get('min') or param.limit.data.get('lower', 'NA')
                
                ws[f'H{current_row}'] = new_ul if new_ul != 'NA' else 'NA'
                ws[f'H{current_row}'].alignment = center_alignment
                ws[f'H{current_row}'].border = thin_border
                
                ws[f'I{current_row}'] = new_ll if new_ll != 'NA' else 'NA'
                ws[f'I{current_row}'].alignment = center_alignment
                ws[f'I{current_row}'].border = thin_border
                
                current_row += 1
            
            # Apply borders to all cells in the used range to ensure consistent borders
            max_row = max(current_row - 1, 9)  # At least 9 rows for the header structure
            for row in range(1, max_row + 1):
                for col in range(1, 10):  # Columns A to I
                    cell = ws.cell(row=row, column=col)
                    if cell.border == Border():  # If no border is set yet
                        cell.border = thin_border
                    # if cell.value is None:
                    #     cell.value = ''
            
            # Add borders to empty cells in column C (gap between left and right sections)
            # for row in range(1, max_row + 1):
            #     ws[f'C{row}'] = ''
            #     ws[f'C{row}'].border = thin_border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 12
            # ws.column_dimensions['C'].width = 3
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            
            # Save workbook
            wb.save(filepath)
            
            # Update UI in main thread
            self.root.after(0, self.excel_export_complete, filepath, None)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, self.excel_export_complete, filepath, str(e))

    def excel_export_complete(self, filepath: str, error: Optional[str]):
        """Called when Excel export is complete"""
        self.export_button.config(state="normal", text="📊 Export to Excel")
        
        if error:
            self.status_label.config(text=f"Excel export error: {error}", fg=MaterialColors.ERROR)
            messagebox.showerror("Error", f"Error exporting Excel file: {error}")
        else:
            self.status_label.config(text="Excel export successful!", fg=MaterialColors.SUCCESS)
            messagebox.showinfo("Success", f"Results exported to file:\n{filepath}")

    def run(self):
        """Chạy ứng dụng"""
        self.root.mainloop()


def main():
    """Chạy GUI"""
    app = CSVComparatorGUI()
    app.run()


if __name__ == "__main__":
    main()
